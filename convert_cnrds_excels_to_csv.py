from __future__ import annotations

import csv
import re
import traceback
import warnings
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(r"c:\Users\21288\Desktop\DACHUANG\dachuang\CNRDS专利数据包")
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
OVERWRITE = False
LOG_PATH = ROOT.parent / 'convert_cnrds_excels_to_csv_progress.log'

warnings.filterwarnings(
    'ignore',
    message='Workbook contains no default style, apply openpyxl\'s default',
)


def safe_name(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', '_', name).strip()
    return cleaned or 'Sheet'


def iter_excel_files(root: Path) -> Iterable[Path]:
    return sorted(
        path for path in root.rglob('*') if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES
    )


def log(message: str) -> None:
    print(message, flush=True)
    with LOG_PATH.open('a', encoding='utf-8') as f:
        f.write(message + '\n')


def workbook_to_csvs(path: Path) -> tuple[list[Path], list[Path]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    output_paths: list[Path] = []
    skipped_paths: list[Path] = []
    try:
        multiple_sheets = len(workbook.sheetnames) > 1
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if multiple_sheets:
                output_path = path.with_name(f"{path.stem}__{safe_name(sheet_name)}.csv")
            else:
                output_path = path.with_suffix('.csv')

            if output_path.exists() and not OVERWRITE:
                skipped_paths.append(output_path)
                continue

            with output_path.open('w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(['' if value is None else value for value in row])

            output_paths.append(output_path)
    finally:
        workbook.close()
    return output_paths, skipped_paths


def main() -> None:
    LOG_PATH.write_text('', encoding='utf-8')
    files = list(iter_excel_files(ROOT))
    log(f'Found {len(files)} Excel files under {ROOT}')
    converted = 0
    produced = 0
    skipped = 0
    failures: list[tuple[Path, str]] = []

    for file_path in files:
        try:
            csv_paths, skipped_paths = workbook_to_csvs(file_path)
            converted += 1
            produced += len(csv_paths)
            skipped += len(skipped_paths)
            if csv_paths:
                log(f'OK   {file_path} -> {len(csv_paths)} csv')
            else:
                log(f'SKIP {file_path} -> already converted')
        except Exception as exc:  # noqa: BLE001
            failures.append((file_path, str(exc)))
            log(f'FAIL {file_path} -> {exc}')

    log('-' * 80)
    log(f'Converted workbooks: {converted}')
    log(f'Created CSV files:   {produced}')
    log(f'Skipped CSV files:   {skipped}')
    log(f'Failures:            {len(failures)}')
    if failures:
        log('Failed files:')
        for path, message in failures:
            log(f'  - {path}: {message}')


if __name__ == '__main__':
    try:
        main()
    except Exception:  # noqa: BLE001
        log('FATAL ERROR')
        log(traceback.format_exc())
        raise
